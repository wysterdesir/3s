"""Extract the ExerciseAnimatic catalogue into one JSON the library generator reads.

Two sources are joined here and nowhere else:
  1. The delivered folder tree — the ONLY 100%-reliable classification signal.
     `HD 720p LOWEST FILE SIZE/<Muscle group>/<name>.mp4`, plus the parallel
     `1200+ GREEN SCREEN VIDEOS/` tree which we prefer when a keyed version of
     the same exercise exists.
  2. `1500+ exercise data.xlsx` — instructions, tips, muscles, equipment. Only
     ~63% filled, so every field it supplies must have a fallback.

    py tools/extract-catalogue.py

Writes tools/catalogues/catalogue.json. Run this on a machine with the bundle
mounted; the generator downstream never touches the Dropbox folder.
"""
import json
import os
import re
import sys

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

BUNDLE = r"C:\Users\wyste\Dropbox\ULTIMATE BUNDLE MASTER FOLDER 4K+1080p+ILLUSTRATIONS+EXERCISE CATALOG"
SD = "HD 720p LOWEST FILE SIZE"
GS = "1200+ GREEN SCREEN VIDEOS"
XLSX = "1500+ exercise data.xlsx"
OUT = os.path.join(os.path.dirname(__file__), "catalogues", "catalogue.json")


def norm(name):
    """Join key: lowercase, no extension, no gender tag, punctuation flattened.

    Their filenames and their spreadsheet disagree on case, on `_Male`/`_female`,
    and on stray whitespace, so both sides get squashed the same way before they
    are compared."""
    s = re.sub(r"\.mp4$", "", name, flags=re.I)
    s = re.sub(r"[_\s-]*(male|female)$", "", s.strip(), flags=re.I)
    s = re.sub(r"[^a-z0-9]+", " ", s.lower()).strip()
    return s


def gender(name):
    m = re.search(r"[_\s-]*(male|female)$", re.sub(r"\.mp4$", "", name, flags=re.I), flags=re.I)
    return m.group(1).lower() if m else ""


def scan(root):
    """folder -> list of filenames, for one resolution tree."""
    out = {}
    base = os.path.join(BUNDLE, root)
    for folder in sorted(os.listdir(base)):
        d = os.path.join(base, folder)
        if not os.path.isdir(d):
            continue
        out[folder] = sorted(f for f in os.listdir(d) if f.lower().endswith(".mp4"))
    return out


def load_sheet():
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(BUNDLE, XLSX), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    next(rows)                                    # header
    by_key = {}
    for r in rows:
        cat, name, steps, tips, prim, sec, equip = (list(r) + [None] * 7)[:7]
        if not name:
            continue
        by_key[norm(str(name))] = {
            "category": (cat or "").strip(),
            "sheetName": str(name).strip(),
            "steps": (steps or "").strip(),
            "tips": (tips or "").strip(),
            "primary": (prim or "").strip(),
            "secondary": (sec or "").strip(),
            "equipment": (equip or "").strip(),
        }
    return by_key


def main():
    sd = scan(SD)
    gs = scan(GS)
    sheet = load_sheet()

    # Green-screen index. Keyed clips drop onto the dark theme without a card, so
    # they win wherever one exists — but only for the SAME model. Indexing on the
    # name alone would hand the male exercise a female green-screen clip and the
    # app would swap cast mid-session.
    gs_index = {}
    for folder, files in gs.items():
        for f in files:
            gs_index.setdefault((norm(f), gender(f)), (folder, f))

    items, matched = [], 0
    for folder, files in sd.items():
        for f in files:
            key = norm(f)
            meta = sheet.get(key)
            if meta:
                matched += 1
            g = gs_index.get((key, gender(f)))
            items.append({
                "key": key,
                "file": f,
                "folder": folder,
                "gender": gender(f),
                "green": {"folder": g[0], "file": g[1]} if g else None,
                "category": meta["category"] if meta else "",
                "steps": meta["steps"] if meta else "",
                "tips": meta["tips"] if meta else "",
                "primary": meta["primary"] if meta else "",
                "secondary": meta["secondary"] if meta else "",
                "equipment": meta["equipment"] if meta else "",
            })

    items.sort(key=lambda it: (it["folder"], it["file"]))
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as fh:
        json.dump({
            "source": {"bundle": os.path.basename(BUNDLE), "tree": SD, "sheet": XLSX},
            "counts": {
                "files": len(items),
                "greenScreen": sum(1 for i in items if i["green"]),
                "sheetRows": len(sheet),
                "sheetMatched": matched,
                "distinctKeys": len({i["key"] for i in items}),
            },
            "items": items,
        }, fh, indent=1)

    print(f"{len(items)} clips across {len(sd)} folders -> {OUT}")
    print(f"  green-screen version available : {sum(1 for i in items if i['green'])}")
    print(f"  joined to the spreadsheet      : {matched} of {len(items)} ({matched * 100 // len(items)}%)")
    print(f"  distinct exercises (gender-merged): {len({i['key'] for i in items})}")
    for field in ("equipment", "primary", "tips", "steps", "category"):
        n = sum(1 for i in items if i[field])
        print(f"  {field:<10} filled: {n:>5} ({n * 100 // len(items)}%)")


if __name__ == "__main__":
    main()
